import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

def auto_adjust_column_width(ws):
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2

def clean_failed_logins(logins):
    cleaned = [line for line in logins if line.strip() and not line.strip().startswith("btmp begins")]
    seen = set()
    unique_logins = []
    for login in cleaned:
        if login not in seen:
            seen.add(login)
            unique_logins.append(login)
    return unique_logins

# Load server data from file
file_path = Path("servers_info.txt")
with open(file_path, "r", encoding="utf-8") as file:
    data = file.read()

# Split by server blocks
servers = re.split(r"=== Server: (.*?) ===", data)[1:]

# After splitting by server blocks
print(f"Number of server blocks found: {len(servers)}")

wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Server List"
ws_summary.append(["IP Address", "Hostname"])

bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Keep track of server sheet names for hyperlinks
server_sheet_names = []

# Define manual servers if you have old OS with old python version (add your IP: Name pairs here)
manual_servers = {
    "IP": "Name",
    # Add more as needed
}

for i in range(0, len(servers), 2):
    ip = servers[i].strip()
    content = servers[i + 1].strip()
    print(f"\nProcessing server: {ip}")
    hostname = re.search(r"Hostname:\s*(.*)", content).group(1) if re.search(r"Hostname:\s*(.*)", content) else "Unknown"
    # Create the server sheet first so we know its name
    ws = wb.create_sheet(title=ip)
    server_sheet_names.append((ip, hostname, ws.title))

    # Disk Usage
    disk_start = next((idx for idx, l in enumerate(content.splitlines()) if l.strip() == "Disk Usage:"), None)
    if disk_start is not None:
        disk_lines = content.splitlines()[disk_start+1:]
        disk_header = disk_lines[0].split()
        disk_rows = []
        for line in disk_lines[1:]:
            if not line.strip() or line.startswith("Mem:") or line.startswith("btmp begins") or line.startswith("=== End of"):
                break
            parts = line.split()
            if len(parts) >= 6:
                disk_rows.append(parts[:6])
        print(f"  Disk header: {disk_header}")
        print(f"  Disk rows: {len(disk_rows)}")
        
        # Write disk data to sheet
        ws.cell(row=2, column=1, value="Disk Usage").font = bold_font
        for c, col in enumerate(disk_header, start=1):
            cell = ws.cell(row=3, column=c, value=col)
            cell.font = bold_font
            cell.border = thin_border
        for r, row in enumerate(disk_rows, start=4):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.border = thin_border
        
        # Add border around disk section
        for row in range(2, len(disk_rows) + 4):
            for col in range(1, len(disk_header) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

    # Memory and Swap Usage
    mem_start = next((idx for idx, l in enumerate(content.splitlines()) if l.strip() == "Memory and Swap Usage:"), None)
    if mem_start is not None:
        mem_lines = content.splitlines()[mem_start+1:]
        mem_header = mem_lines[0].split()
        mem_rows = []
        for line in mem_lines[1:]:
            if line.strip().startswith("Mem:") or line.strip().startswith("Swap:"):
                parts = line.split()
                # Pad or trim to match header length
                if len(parts) < len(mem_header):
                    parts += [''] * (len(mem_header) - len(parts))
                elif len(parts) > len(mem_header):
                    parts = parts[:len(mem_header)]
                mem_rows.append(parts)
        print(f"  Memory header: {mem_header}")
        print(f"  Memory rows: {len(mem_rows)}")
        
        # Write memory data to sheet
        mem_col = len(disk_header) + 3
        ws.cell(row=2, column=mem_col, value="Memory and Swap Usage").font = bold_font
        for c, col in enumerate(mem_header, start=mem_col):
            cell = ws.cell(row=3, column=c, value=col)
            cell.font = bold_font
            cell.border = thin_border
        for r, row in enumerate(mem_rows, start=4):
            for c, value in enumerate(row, start=mem_col):
                cell = ws.cell(row=r, column=c, value=value)
                cell.border = thin_border
        
        # Add border around memory section
        for row in range(2, len(mem_rows) + 4):
            for col in range(mem_col, mem_col + len(mem_header)):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

    # Failed Logins
    failed_start = next((idx for idx, l in enumerate(content.splitlines()) if l.strip() == "Failed Logins:"), None)
    if failed_start is not None:
        failed_lines = content.splitlines()[failed_start+1:]
        # Filter out 'btmp begins', 'End of', and empty lines
        failed_logins = [line for line in failed_lines if line.strip() and not line.strip().startswith("btmp begins") and not line.strip().startswith("=== End of")]
        print(f"  Failed logins: {len(failed_logins)}")
        
        # Write failed logins to sheet (under memory section)
        failed_row = len(mem_rows) + 6
        ws.cell(row=failed_row, column=mem_col, value="Failed Logins").font = bold_font
        headers = ["User", "Service", "IP Address", "Date/Time"]
        for c, col in enumerate(headers, start=mem_col):
            cell = ws.cell(row=failed_row+1, column=c, value=col)
            cell.font = bold_font
            cell.border = thin_border
        
        for r, login in enumerate(failed_logins, start=failed_row+2):
            parts = login.split()
            if len(parts) >= 4:
                # Reorder parts to match the headers: User, Service, IP Address, Date/Time
                user = parts[0]
                service = parts[1]
                ip = parts[2]
                datetime = ' '.join(parts[3:])  # Combine remaining parts for date/time
                
                ws.cell(row=r, column=mem_col, value=user).border = thin_border
                ws.cell(row=r, column=mem_col+1, value=service).border = thin_border
                ws.cell(row=r, column=mem_col+2, value=ip).border = thin_border
                ws.cell(row=r, column=mem_col+3, value=datetime).border = thin_border
        
        # Add border around failed logins section
        for row in range(failed_row, failed_row + len(failed_logins) + 2):
            for col in range(mem_col, mem_col + len(headers)):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

    # Add back to main page link at the top
    ws.cell(row=1, column=1).value = "Back to Main Page"
    ws.cell(row=1, column=1).hyperlink = "#'Server List'!A1"
    ws.cell(row=1, column=1).style = "Hyperlink"
    
    # Auto-adjust column widths
    auto_adjust_column_width(ws)

# Now fill the summary sheet with hyperlinks
for idx, (ip, hostname, sheet_name) in enumerate(server_sheet_names, start=2):
    ws_summary.cell(row=idx, column=1).value = ip
    ws_summary.cell(row=idx, column=1).hyperlink = f"#{sheet_name}!A1"
    ws_summary.cell(row=idx, column=1).style = "Hyperlink"
    ws_summary.cell(row=idx, column=2).value = hostname
    ws_summary.cell(row=idx, column=2).hyperlink = f"#{sheet_name}!A1"
    ws_summary.cell(row=idx, column=2).style = "Hyperlink"

# After filling the summary sheet with auto servers, add manual servers
manual_start_row = len(server_sheet_names) + 2
for idx, (ip, hostname) in enumerate(manual_servers.items(), start=manual_start_row):
    sheet_name = ip
    ws = wb.create_sheet(title=sheet_name)
    ws_summary.cell(row=idx, column=1).value = ip
    ws_summary.cell(row=idx, column=1).hyperlink = f"#{sheet_name}!A1"
    ws_summary.cell(row=idx, column=1).style = "Hyperlink"
    ws_summary.cell(row=idx, column=2).value = hostname
    ws_summary.cell(row=idx, column=2).hyperlink = f"#{sheet_name}!A1"
    ws_summary.cell(row=idx, column=2).style = "Hyperlink"

    # Add back to main page link at the top
    ws.cell(row=1, column=1).value = "Back to Main Page"
    ws.cell(row=1, column=1).hyperlink = "#'Server List'!A1"
    ws.cell(row=1, column=1).style = "Hyperlink"
    start_row = 2
    # Disk Usage section
    ws.cell(row=start_row, column=1, value="Disk Usage").font = bold_font
    disk_headers = ["Filesystem", "Size", "Used", "Avail", "Use%", "Mounted-on"]
    for c, col in enumerate(disk_headers, start=1):
        cell = ws.cell(row=start_row+1, column=c, value=col)
        cell.font = bold_font
        cell.border = thin_border
    # Memory and Swap Usage section
    mem_col = len(disk_headers) + 3
    ws.cell(row=start_row, column=mem_col, value="Memory and Swap Usage").font = bold_font
    mem_headers = ["Type", "total", "used", "free", "shared", "buff/cache", "available"]
    for c, col in enumerate(mem_headers, start=mem_col):
        cell = ws.cell(row=start_row+1, column=c, value=col)
        cell.font = bold_font
        cell.border = thin_border
    mem_last_row = start_row+1
    # Failed Logins section
    failed_row = mem_last_row + 2
    ws.cell(row=failed_row, column=mem_col, value="Failed Logins").font = bold_font
    ws.cell(row=failed_row, column=mem_col).border = thin_border
    ws.cell(row=failed_row+1, column=mem_col, value="User").border = thin_border
    ws.cell(row=failed_row+1, column=mem_col+1, value="Service").border = thin_border
    ws.cell(row=failed_row+1, column=mem_col+2, value="IP Address").border = thin_border
    ws.cell(row=failed_row+1, column=mem_col+3, value="Date/Time").border = thin_border
    ws.freeze_panes = "A2"
    auto_adjust_column_width(ws)

output_path = Path("server-report.xlsx")
wb.save(output_path)
print(f"Excel report saved to: {output_path}")
